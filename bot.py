import asyncio
import logging
from collections import defaultdict
from datetime import date, datetime, timedelta

from aiogram import Bot, Dispatcher, F, types
from aiogram.exceptions import TelegramForbiddenError
from aiogram.filters import Command, StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import (
    InlineKeyboardMarkup,
    ReplyKeyboardMarkup,
    KeyboardButton,
    ReplyKeyboardRemove,
)
from aiogram.utils.keyboard import InlineKeyboardBuilder, ReplyKeyboardBuilder

import database as db
from config import BOT_TOKEN, MASTER_CODE, ADMIN_ID
from storage import SQLiteStorage
from middleware import MasterContextMiddleware, BlockedUserMiddleware

from logging.handlers import RotatingFileHandler

_log_formatter = logging.Formatter("%(asctime)s [%(levelname)s] %(name)s: %(message)s")

_file_handler = RotatingFileHandler(
    "bot.log", maxBytes=5 * 1024 * 1024, backupCount=3, encoding="utf-8"
)
_file_handler.setFormatter(_log_formatter)

_console_handler = logging.StreamHandler()
_console_handler.setFormatter(_log_formatter)

logging.basicConfig(level=logging.INFO, handlers=[_console_handler, _file_handler])
logger = logging.getLogger(__name__)

storage = SQLiteStorage("salon_bot.db")
bot = Bot(token=BOT_TOKEN)
dp = Dispatcher(storage=storage)
dp.message.middleware(BlockedUserMiddleware())
dp.callback_query.middleware(BlockedUserMiddleware())
dp.message.middleware(MasterContextMiddleware())
dp.callback_query.middleware(MasterContextMiddleware())


# ─────────────────────────── States ───────────────────────────

class MasterAuth(StatesGroup):
    entering_password = State()

class MasterReg(StatesGroup):
    entering_code = State()
    name = State()
    service_type = State()
    setting_password = State()
    confirming_password = State()

class MasterSlots(StatesGroup):
    selecting_date = State()
    selecting_times = State()

class MasterChangePwd(StatesGroup):
    old_password = State()
    new_password = State()
    confirm_password = State()

class MasterPriceSetting(StatesGroup):
    entering_price = State()

class MasterUpdateInfo(StatesGroup):
    name         = State()
    service_type = State()

class MasterBroadcast(StatesGroup):
    entering_text = State()

class AdminAddMaster(StatesGroup):
    name = State()
    service_type = State()

class AdminSchedule(StatesGroup):
    selecting_master = State()
    selecting_action = State()
    setting_hours    = State()
    selecting_day_off = State()

class ClientBooking(StatesGroup):
    selecting_service = State()
    selecting_master = State()
    entering_name = State()
    entering_phone = State()
    selecting_slot = State()

# ─────────────────────────── Constants ────────────────────────

SERVICES = {
    "tattoo":  "💉 Тату",
    "makeup":  "💄 Макияж",
    "nails":   "💅 Маникюр",
    "hair":    "✂️ Парикмахер",
    "massage": "💆 Массаж",
    "other":   "🌟 Другое",
}
DAYS_RU = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]

def is_admin(user_id: int) -> bool:
    return ADMIN_ID != 0 and user_id == ADMIN_ID

# ─────────────────────────── Keyboards ────────────────────────

def kb_start() -> InlineKeyboardMarkup:
    b = InlineKeyboardBuilder()
    b.button(text="👨‍🎨 Я мастер",  callback_data="role:master")
    b.button(text="👤 Я клиент",   callback_data="role:client")
    b.button(text="📋 Мои записи", callback_data="my_bookings")
    b.adjust(2, 1)
    return b.as_markup()

def kb_services(prefix: str) -> InlineKeyboardMarkup:
    b = InlineKeyboardBuilder()
    for key, label in SERVICES.items():
        b.button(text=label, callback_data=f"{prefix}:{key}")
    b.adjust(2)
    return b.as_markup()

def kb_master_menu() -> ReplyKeyboardMarkup:
    b = ReplyKeyboardBuilder()
    b.button(text="📅 Добавить слоты")
    b.button(text="📋 Мой график")
    b.button(text="🕐 История")
    b.button(text="👥 Клиенты")
    b.button(text="📊 Статистика")
    b.button(text="📢 Рассылка")
    b.button(text="📥 Экспорт Excel")
    b.button(text="🔄 Шаблон")
    b.button(text="⚙️ Настройки")
    b.adjust(2)
    return b.as_markup(resize_keyboard=True)

def kb_admin_menu() -> InlineKeyboardMarkup:
    b = InlineKeyboardBuilder()
    b.button(text="➕ Добавить мастера",       callback_data="adm_add")
    b.button(text="👨‍🎨 Управление мастерами",  callback_data="adm_masters")
    b.button(text="📅 Расписание мастеров",    callback_data="adm_schedule")
    b.button(text="⏳ Ожидающие активацию",    callback_data="adm_pending")
    b.button(text="📊 Отчёт за месяц",         callback_data="adm_report")
    b.adjust(1)
    return b.as_markup()

MONTHS_RU = [
    "", "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
    "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь",
]

def kb_months() -> InlineKeyboardMarkup:
    b = InlineKeyboardBuilder()
    today = date.today()
    seen = set()
    for i in range(90):
        d = today + timedelta(days=i)
        key = f"{d.year}-{d.month:02d}"
        if key not in seen:
            seen.add(key)
            b.button(
                text=f"📅 {MONTHS_RU[d.month]} {d.year}",
                callback_data=f"slot_month:{key}",
            )
    b.button(text="❌ Отмена", callback_data="cancel")
    b.adjust(1)
    return b.as_markup()

def kb_days_in_month(year: int, month: int) -> InlineKeyboardMarkup:
    import calendar
    b = InlineKeyboardBuilder()
    today = date.today()
    _, days_in_month = calendar.monthrange(year, month)
    for day in range(1, days_in_month + 1):
        d = date(year, month, day)
        if d < today:
            continue
        b.button(
            text=f"{d.strftime('%d')} {DAYS_RU[d.weekday()]}",
            callback_data=f"slot_date:{d.isoformat()}",
        )
    b.button(text="🔙 Назад", callback_data="slot_back_months")
    b.button(text="❌ Отмена", callback_data="cancel")
    b.adjust(4)
    return b.as_markup()

def kb_times(selected: list[str], sel_date: str) -> InlineKeyboardMarkup:
    b = InlineKeyboardBuilder()
    for hour in range(9, 21):
        for minute in (0, 30):
            t = f"{hour:02d}:{minute:02d}"
            mark = "✅ " if t in selected else ""
            b.button(text=f"{mark}{t}", callback_data=f"slot_time:{sel_date}:{hour:02d}:{minute:02d}")
    b.button(text="💾 Сохранить", callback_data=f"save_slots:{sel_date}")
    b.button(text="❌ Отмена",    callback_data="cancel")
    b.adjust(4)
    return b.as_markup()

def kb_slots(slots, price: int = 0) -> InlineKeyboardMarkup:
    b = InlineKeyboardBuilder()
    grouped: dict = defaultdict(list)
    for s in slots:
        grouped[s["date"]].append(s)
    for date_str in sorted(grouped):
        d = date.fromisoformat(date_str)
        for s in sorted(grouped[date_str], key=lambda x: x["time"]):
            price_str = f" · {price} ₽" if price else ""
            b.button(
                text=f"📅 {d.strftime('%d.%m.%Y')} ({DAYS_RU[d.weekday()]}) {s['time']}{price_str}",
                callback_data=f"book:{s['id']}",
            )
    b.button(text="❌ Отмена", callback_data="cancel")
    b.adjust(1)
    return b.as_markup()

def build_schedule_message(slots, offset: int = 0) -> tuple[str, InlineKeyboardMarkup]:
    """Returns (text, keyboard). Shows 7 days with prev/next navigation."""
    start = date.today() + timedelta(days=offset)
    end   = start + timedelta(days=7)

    grouped: dict = defaultdict(list)
    for s in slots:
        grouped[s["date"]].append(s)

    lines = [f"📋 *График: {start.strftime('%d.%m')} – {(end - timedelta(days=1)).strftime('%d.%m')}*\n"]
    rows: list = []

    if not grouped:
        lines.append("_Слотов на эту неделю нет_")
    else:
        for date_str in sorted(grouped):
            d = date.fromisoformat(date_str)
            day_slots = sorted(grouped[date_str], key=lambda x: x["time"])
            free_times = [s["time"] for s in day_slots if not s["is_booked"]]
            booked     = [s for s in day_slots if s["is_booked"] and s["client_name"]]

            lines.append(f"📅 *{d.strftime('%d.%m')} ({DAYS_RU[d.weekday()]})*")
            if free_times:
                lines.append("🟢 " + " · ".join(free_times))
                rows.append([types.InlineKeyboardButton(
                    text=f"🗑 Убрать свободные {d.strftime('%d.%m')}",
                    callback_data=f"clear_day:{date_str}",
                )])
            for s in booked:
                if s["pending"]:
                    lines.append(f"⏳ {s['time']} — *{s['client_name']}* _(ожидает подтв.)_")
                    rows.append([types.InlineKeyboardButton(
                        text=f"✅ Подтвердить {s['time']} {s['client_name'][:12]}",
                        callback_data=f"confirm_book:{s['booking_id']}",
                    )])
                    rows.append([types.InlineKeyboardButton(
                        text=f"❌ Отклонить {s['time']}",
                        callback_data=f"reject_book:{s['booking_id']}",
                    )])
                else:
                    lines.append(f"🔴 {s['time']} — *{s['client_name']}*")
                    rows.append([types.InlineKeyboardButton(
                        text=f"❌ {s['time']} {s['client_name'][:14]} — отменить (уведомить)",
                        callback_data=f"mcb:{s['id']}",
                    )])
                    rows.append([types.InlineKeyboardButton(
                        text=f"🗑 {s['time']} — удалить без уведомления",
                        callback_data=f"manualdel:{s['id']}",
                    )])
            lines.append("")

    nav = []
    if offset > 0:
        nav.append(types.InlineKeyboardButton(
            text="◀️ Назад", callback_data=f"sched_page:{offset - 7}",
        ))
    nav.append(types.InlineKeyboardButton(
        text="▶️ Вперёд", callback_data=f"sched_page:{offset + 7}",
    ))
    rows.append(nav)

    return "\n".join(lines), InlineKeyboardMarkup(inline_keyboard=rows)

def kb_client_bookings(bookings) -> InlineKeyboardMarkup:
    b = InlineKeyboardBuilder()
    now = datetime.now()
    for bk in bookings:
        appt_dt = datetime.strptime(f"{bk['date']} {bk['time']}", "%Y-%m-%d %H:%M")
        if (appt_dt - now).total_seconds() / 60 > 120:
            d = date.fromisoformat(bk["date"])
            b.button(
                text=f"❌ Отменить {d.strftime('%d.%m')} {bk['time']} — {bk['master_name']}",
                callback_data=f"ccb:{bk['slot_id']}",
            )
    b.adjust(1)
    return b.as_markup()

def kb_cancel_confirm(callback_yes: str) -> InlineKeyboardMarkup:
    b = InlineKeyboardBuilder()
    b.button(text="✅ Да, отменить", callback_data=callback_yes)
    b.button(text="❌ Нет",          callback_data="confirm_no")
    b.adjust(2)
    return b.as_markup()

def kb_cancel_text() -> ReplyKeyboardMarkup:
    """Single cancel button for text-input states."""
    b = ReplyKeyboardBuilder()
    b.button(text="❌ Отмена")
    return b.as_markup(resize_keyboard=True, one_time_keyboard=True)

def kb_masters_manage(masters) -> InlineKeyboardMarkup:
    b = InlineKeyboardBuilder()
    for m in masters:
        status = "✅" if m["is_active"] else "❌"
        b.button(
            text=f"{status} {m['name']} — {SERVICES.get(m['service_type'], '')}",
            callback_data=f"adm_toggle:{m['id']}",
        )
    b.button(text="🔙 Назад", callback_data="adm_back")
    b.adjust(1)
    return b.as_markup()

# ─────────────────────────── Reminders background ─────────────

async def _send_reminders():
    bookings = await db.get_pending_reminders()
    now = datetime.now()
    current_hour = now.hour  # for 8h reminder window check
    for b in bookings:
        appt_dt = datetime.strptime(f"{b['date']} {b['time']}", "%Y-%m-%d %H:%M")
        mins = (appt_dt - now).total_seconds() / 60
        d_fmt = date.fromisoformat(b["date"]).strftime("%d.%m.%Y")
        svc = SERVICES.get(b["service_type"], "")
        if 1380 <= mins <= 1500 and not b["reminder_24h_sent"]:
            try:
                await bot.send_message(
                    b["client_telegram_id"],
                    f"⏰ *Напоминание!*\nЗавтра у вас запись:\n"
                    f"👨‍🎨 {b['master_name']} — {svc}\n📅 {d_fmt}\n🕐 {b['time']}",
                    parse_mode="Markdown",
                )
                await db.mark_reminder_sent(b["id"], "24h")
            except TelegramForbiddenError:
                cancelled = await db.auto_cancel_client(b["client_telegram_id"])
                logger.info(f"Клиент {b['client_telegram_id']} заблокировал бот — отменено {cancelled} записей")
            except Exception as e:
                logger.warning(f"24h reminder failed: {e}")
        elif 450 <= mins <= 510 and not b["reminder_8h_sent"] and 7 <= current_hour < 22:
            try:
                await bot.send_message(
                    b["client_telegram_id"],
                    f"⏰ *Напоминание — сегодня запись!*\n\n"
                    f"👨‍🎨 {b['master_name']} — {svc}\n📅 {d_fmt}\n🕐 {b['time']}\n\n"
                    "_Если не сможете прийти — отмените через /mybookings_",
                    parse_mode="Markdown",
                )
                await db.mark_reminder_sent(b["id"], "8h")
            except TelegramForbiddenError:
                cancelled = await db.auto_cancel_client(b["client_telegram_id"])
                logger.info(f"Клиент {b['client_telegram_id']} заблокировал бот — отменено {cancelled} записей")
            except Exception as e:
                logger.warning(f"8h reminder failed: {e}")
        elif 45 <= mins <= 75 and not b["reminder_1h_sent"]:
            try:
                await bot.send_message(
                    b["client_telegram_id"],
                    f"⏰ *Через час ваша запись!*\n"
                    f"👨‍🎨 {b['master_name']} — {svc}\n📅 {d_fmt}\n🕐 {b['time']}",
                    parse_mode="Markdown",
                )
                await db.mark_reminder_sent(b["id"], "1h")
            except TelegramForbiddenError:
                cancelled = await db.auto_cancel_client(b["client_telegram_id"])
                logger.info(f"Клиент {b['client_telegram_id']} заблокировал бот — отменено {cancelled} записей")
            except Exception as e:
                logger.warning(f"1h reminder failed: {e}")

async def _send_review_requests():
    """Ask clients to rate their visit (30-180 min after appointment)."""
    now = datetime.now()
    targets = await db.get_review_targets()
    for r in targets:
        appt_dt = datetime.strptime(f"{r['date']} {r['time']}", "%Y-%m-%d %H:%M")
        mins_past = (now - appt_dt).total_seconds() / 60
        if 30 <= mins_past <= 180:
            try:
                brev = InlineKeyboardBuilder()
                for score in range(1, 6):
                    brev.button(text="⭐" * score, callback_data=f"review:{r['id']}:{score}")
                brev.adjust(5)
                await bot.send_message(
                    r["client_telegram_id"],
                    f"🌟 Как прошёл ваш визит к *{r['master_name']}*?\n\nОцените качество услуги:",
                    parse_mode="Markdown",
                    reply_markup=brev.as_markup(),
                )
                await db.mark_review_asked(r["id"])
            except TelegramForbiddenError:
                await db.mark_review_asked(r["id"])
            except Exception as e:
                logger.warning(f"Review request failed for booking {r['id']}: {e}")


async def reminder_loop():
    while True:
        try:
            await _send_reminders()
            await _send_review_requests()
        except Exception as e:
            logger.error(f"Reminder loop: {e}")
        await asyncio.sleep(300)


async def cleanup_loop():
    """Runs once per day: removes past free slots and old bookings."""
    while True:
        await asyncio.sleep(86400)  # 24 hours
        try:
            result = await db.cleanup_old_data()
            logger.info(
                f"Cleanup: удалено свободных слотов={result['free_slots']}, "
                f"старых записей={result['old_bookings']}, orphan={result['orphan_slots']}"
            )
        except Exception as e:
            logger.error(f"Cleanup loop: {e}")

# ─────────────────────────── /start  /help  /myid ─────────────

@dp.message(Command("start"))
async def cmd_start(message: types.Message, state: FSMContext, master=None):
    current_state = await state.get_state()
    if master is None:
        master = await db.get_master(message.from_user.id)
    if master:
        await db.reset_wrong_attempts(message.from_user.id)
        # Already authenticated (has menu open) — just show menu again
        if current_state is None and master["password_hash"]:
            await message.answer(
                f"👋 Привет, *{master['name']}*! Главное меню:",
                parse_mode="Markdown", reply_markup=kb_master_menu(),
            )
            return
        await state.clear()
        if not master["password_hash"]:
            await state.set_state(MasterReg.setting_password)
            await state.update_data(name=master["name"], service_type=master["service_type"])
            await message.answer(
                f"👋 Привет, *{master['name']}*!\n\n🔐 Установите пароль для входа:\n_(минимум 4 символа)_",
                parse_mode="Markdown", reply_markup=ReplyKeyboardRemove(),
            )
        else:
            await state.set_state(MasterAuth.entering_password)
            await message.answer(
                f"🔐 Добро пожаловать, *{master['name']}*!\n\nВведите ваш пароль:\n_(или /cancel для отмены)_",
                parse_mode="Markdown", reply_markup=ReplyKeyboardRemove(),
            )
    else:
        await state.clear()
        text = "👋 Добро пожаловать!\n\nВы мастер или клиент?"
        if is_admin(message.from_user.id):
            text += "\n\n🔑 Вы вошли как *администратор*. Используйте /admin"
        await message.answer(text, parse_mode="Markdown", reply_markup=kb_start())


@dp.message(Command("myid"))
async def cmd_myid(message: types.Message):
    await message.answer(
        f"Ваш Telegram ID: <code>{message.from_user.id}</code>\n\n"
        "Вставьте это число в <code>config.py</code> — <code>ADMIN_ID = ...</code>",
        parse_mode="HTML",
    )


@dp.message(Command("mybookings"))
async def cmd_mybookings(message: types.Message):
    bookings = await db.get_client_bookings(message.from_user.id)
    if not bookings:
        await message.answer(
            "📋 У вас нет предстоящих записей.\n\nЗапишитесь через /start → «👤 Я клиент»."
        )
        return
    now = datetime.now()
    text = "📋 *Ваши предстоящие записи:*\n\n"
    has_cancel = False
    for i, bk in enumerate(bookings, 1):
        d = date.fromisoformat(bk["date"])
        appt_dt = datetime.strptime(f"{bk['date']} {bk['time']}", "%Y-%m-%d %H:%M")
        status_tag = " ⏳ _ожидает подтв._" if bk.get("pending") else ""
        text += (
            f"*{i}.* 📅 {d.strftime('%d.%m.%Y')} ({DAYS_RU[d.weekday()]}) в {bk['time']}{status_tag}\n"
            f"   👨‍🎨 {bk['master_name']} — {SERVICES.get(bk['service_type'], '')}\n\n"
        )
        if (appt_dt - now).total_seconds() / 60 > 120:
            has_cancel = True
    if has_cancel:
        text += "_Нажмите кнопку ниже, чтобы отменить запись (доступно за 2+ часа до визита):_"
    await message.answer(text, parse_mode="Markdown", reply_markup=kb_client_bookings(bookings))


@dp.message(Command("cancel"))
async def cmd_cancel(message: types.Message, state: FSMContext):
    await state.clear()
    master = await db.get_master(message.from_user.id)
    if master and master["password_hash"]:
        await message.answer("↩️ Действие отменено.", reply_markup=kb_master_menu())
    else:
        await message.answer("↩️ Действие отменено.", reply_markup=ReplyKeyboardRemove())
        await message.answer("Главное меню:", reply_markup=kb_start())


@dp.message(F.text == "❌ Отмена")
async def cancel_text_button(message: types.Message, state: FSMContext):
    await state.clear()
    master = await db.get_master(message.from_user.id)
    if master and master["password_hash"]:
        await message.answer("↩️ Отменено.", reply_markup=kb_master_menu())
    else:
        await message.answer("↩️ Отменено.", reply_markup=ReplyKeyboardRemove())
        await message.answer("Главное меню:", reply_markup=kb_start())


@dp.message(Command("help"))
async def cmd_help(message: types.Message):
    await message.answer(
        "ℹ️ *Команды:*\n/start — главное меню\n/mybookings — мои записи и отмена\n"
        "/cancel — отменить текущее действие\n"
        "/myid — узнать свой Telegram ID\n/admin — панель администратора\n\n"
        "*Мастеру:* добавляйте слоты и следите за графиком.\n"
        "*Клиенту:* выберите услугу, мастера и запишитесь.",
        parse_mode="Markdown",
    )

# ─────────────────────────── ADMIN PANEL ──────────────────────

@dp.message(Command("admin"))
async def cmd_admin(message: types.Message, state: FSMContext):
    await state.clear()
    if ADMIN_ID == 0:
        await message.answer(
            f"⚙️ ADMIN_ID не настроен.\n\n"
            f"Ваш Telegram ID: <code>{message.from_user.id}</code>\n"
            "Вставьте его в <code>config.py</code> — <code>ADMIN_ID = ...</code> — и перезапустите бота.",
            parse_mode="HTML",
        )
        return
    if not is_admin(message.from_user.id):
        return
    await message.answer("🔑 *Панель администратора*", parse_mode="Markdown",
                         reply_markup=kb_admin_menu())


# Добавить мастера
@dp.callback_query(F.data == "adm_add")
async def adm_add(callback: types.CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        return
    await state.set_state(AdminAddMaster.name)
    await callback.message.edit_text("✏️ Введите имя нового мастера:\n_(или /cancel для отмены)_")


@dp.message(StateFilter(AdminAddMaster.name))
async def adm_add_name(message: types.Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    name = message.text.strip()
    if len(name) < 2:
        await message.answer("❌ Слишком короткое имя:")
        return
    await state.update_data(name=name)
    await state.set_state(AdminAddMaster.service_type)
    await message.answer("🎨 Выберите тип услуги нового мастера:", reply_markup=kb_services("adm_svc"))


@dp.callback_query(F.data.startswith("adm_svc:"))
async def adm_add_service(callback: types.CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        return
    service_type = callback.data.split(":")[1]
    data = await state.get_data()
    name = data["name"]
    code = await db.admin_create_master(name, service_type)
    await state.clear()
    await callback.message.edit_text(
        f"✅ *Мастер создан!*\n\n"
        f"👤 Имя: {name}\n"
        f"🎨 Услуга: {SERVICES.get(service_type)}\n\n"
        f"📨 *Код приглашения для мастера:*\n"
        f"`{code}`\n\n"
        "Отправьте этот код мастеру. Мастер входит в бот, нажимает «Я мастер», вводит код, ставит свой пароль.",
        parse_mode="Markdown",
        reply_markup=kb_admin_menu(),
    )


# Управление мастерами (активировать / деактивировать)
@dp.callback_query(F.data == "adm_masters")
async def adm_masters(callback: types.CallbackQuery):
    if not is_admin(callback.from_user.id):
        return
    masters = await db.get_all_masters()
    if not masters:
        await callback.message.edit_text("Мастеров пока нет.", reply_markup=kb_admin_menu())
        return
    await callback.message.edit_text(
        "👨‍🎨 *Список мастеров*\n\n✅ — активен · ❌ — деактивирован\nНажмите, чтобы изменить статус:",
        parse_mode="Markdown",
        reply_markup=kb_masters_manage(masters),
    )


@dp.callback_query(F.data.startswith("adm_toggle:"))
async def adm_toggle(callback: types.CallbackQuery):
    if not is_admin(callback.from_user.id):
        return
    master_id = int(callback.data.split(":")[1])
    new_active = await db.toggle_master_active(master_id)
    status = "восстановлен ✅" if new_active else "деактивирован ❌"
    await callback.answer(f"Мастер {status}", show_alert=True)
    masters = await db.get_all_masters()
    await callback.message.edit_reply_markup(reply_markup=kb_masters_manage(masters))


# Ожидающие активацию (созданные, но ещё не вошедшие)
@dp.callback_query(F.data == "adm_pending")
async def adm_pending(callback: types.CallbackQuery):
    if not is_admin(callback.from_user.id):
        return
    pending = await db.get_pending_invites()
    if not pending:
        await callback.message.edit_text(
            "✅ Нет ожидающих активации мастеров.", reply_markup=kb_admin_menu()
        )
        return
    text = "⏳ *Ожидают активации:*\n\n"
    for p in pending:
        text += f"👤 {p['name']} — {SERVICES.get(p['service_type'], '')}\n🔑 Код: `{p['code']}`\n\n"
    await callback.message.edit_text(text, parse_mode="Markdown", reply_markup=kb_admin_menu())


# Месячный отчёт
@dp.callback_query(F.data == "adm_report")
async def adm_report(callback: types.CallbackQuery):
    if not is_admin(callback.from_user.id):
        return
    report = await db.get_admin_monthly_report()
    totals = report["totals"]
    month_name = date.today().strftime("%B %Y")

    text = (
        f"📅 *Отчёт за {month_name}*\n\n"
        f"📊 *Итого:*\n"
        f"✅ Записей: {totals['total_bookings']}\n"
        f"❌ Отменено: {totals['total_cancelled']}\n"
        f"💰 Доход: {totals['total_revenue']:,} ₽\n\n"
        "👨‍🎨 *По мастерам:*\n"
    )
    for m in report["by_master"]:
        revenue_str = f" · 💰 {m['revenue']:,} ₽" if m["price"] else ""
        text += (
            f"\n*{m['name']}* — {SERVICES.get(m['service_type'], '')}\n"
            f"✅ {m['bookings']} зап. · ❌ {m['cancelled']} отм.{revenue_str}\n"
        )
    await callback.message.edit_text(text, parse_mode="Markdown", reply_markup=kb_admin_menu())


@dp.callback_query(F.data == "adm_back")
async def adm_back(callback: types.CallbackQuery):
    if not is_admin(callback.from_user.id):
        return
    await callback.message.edit_text("🔑 *Панель администратора*",
                                     parse_mode="Markdown", reply_markup=kb_admin_menu())


# ─────────────────────────── Admin: schedule ──────────────────

def kb_masters_for_schedule(masters) -> InlineKeyboardMarkup:
    b = InlineKeyboardBuilder()
    for m in masters:
        b.button(
            text=f"👨‍🎨 {m['name']} — {SERVICES.get(m['service_type'], '')}",
            callback_data=f"adms_m:{m['id']}",
        )
    b.button(text="🔙 Назад", callback_data="adm_back")
    b.adjust(1)
    return b.as_markup()


def kb_schedule_actions(master_id: int, tmpl) -> InlineKeyboardMarkup:
    b = InlineKeyboardBuilder()
    if tmpl:
        sh, sm = tmpl["start_hour"], tmpl["start_min"]
        eh, em = tmpl["end_hour"], tmpl["end_min"]
        b.button(
            text=f"⚙️ Часы: {sh:02d}:{sm:02d} – {eh:02d}:{em:02d} (изменить)",
            callback_data=f"adms_sethours:{master_id}",
        )
    else:
        b.button(text="⚙️ Задать часы работы", callback_data=f"adms_sethours:{master_id}")
    b.button(text="✅ Применить расписание на 90 дней", callback_data=f"adms_apply:{master_id}")
    b.button(text="🚫 Сделать выходной",                callback_data=f"adms_dayoff:{master_id}")
    b.button(text="🔙 Назад",                           callback_data="adm_schedule")
    b.adjust(1)
    return b.as_markup()


def kb_dayoff_dates(master_id: int) -> InlineKeyboardMarkup:
    b = InlineKeyboardBuilder()
    today = date.today()
    for i in range(90):
        d = today + timedelta(days=i)
        b.button(
            text=f"{d.strftime('%d.%m')} ({DAYS_RU[d.weekday()]})",
            callback_data=f"adms_dodate:{master_id}:{d.isoformat()}",
        )
    b.button(text="🔙 Назад", callback_data=f"adms_m:{master_id}")
    b.adjust(3)
    return b.as_markup()


@dp.callback_query(F.data == "adm_schedule")
async def adm_schedule(callback: types.CallbackQuery):
    if not is_admin(callback.from_user.id):
        return
    masters = await db.get_all_masters()
    if not masters:
        await callback.message.edit_text("Мастеров нет.", reply_markup=kb_admin_menu())
        return
    await callback.message.edit_text(
        "📅 *Расписание мастеров*\n\nВыберите мастера:",
        parse_mode="Markdown",
        reply_markup=kb_masters_for_schedule(masters),
    )


@dp.callback_query(F.data.startswith("adms_m:"))
async def adms_master_selected(callback: types.CallbackQuery):
    if not is_admin(callback.from_user.id):
        return
    master_id = int(callback.data.split(":")[1])
    masters = await db.get_all_masters()
    master = next((m for m in masters if m["id"] == master_id), None)
    if not master:
        await callback.answer("Мастер не найден.", show_alert=True)
        return
    tmpl = await db.get_schedule_template(master_id)
    tmpl_text = ""
    if tmpl:
        tmpl_text = (f"\n\n⏰ Текущий шаблон: *{tmpl['start_hour']:02d}:{tmpl['start_min']:02d}"
                     f" – {tmpl['end_hour']:02d}:{tmpl['end_min']:02d}*"
                     f", каждые {tmpl['interval']} мин")
    await callback.message.edit_text(
        f"👨‍🎨 *{master['name']}* — {SERVICES.get(master['service_type'], '')}{tmpl_text}\n\n"
        "Выберите действие:",
        parse_mode="Markdown",
        reply_markup=kb_schedule_actions(master_id, tmpl),
    )


@dp.callback_query(F.data.startswith("adms_sethours:"))
async def adms_set_hours_start(callback: types.CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        return
    master_id = int(callback.data.split(":")[1])
    await state.set_state(AdminSchedule.setting_hours)
    await state.update_data(schedule_master_id=master_id)
    await callback.message.edit_text(
        "⏰ Введите часы работы в формате:\n`ЧЧ:ММ-ЧЧ:ММ`\n\n"
        "Например: `10:00-19:00` или `09:30-18:00`\n_(или /cancel для отмены)_",
        parse_mode="Markdown",
    )


@dp.message(StateFilter(AdminSchedule.setting_hours))
async def adms_set_hours_input(message: types.Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    text = message.text.strip().replace(" ", "")
    try:
        start_str, end_str = text.split("-")
        sh, sm = map(int, start_str.split(":"))
        eh, em = map(int, end_str.split(":"))
        assert 0 <= sh <= 23 and 0 <= sm < 60
        assert 0 <= eh <= 23 and 0 <= em < 60
        assert sh * 60 + sm < eh * 60 + em
    except Exception:
        await message.answer(
            "❌ Неверный формат. Введите как `10:00-19:00`:",
            parse_mode="Markdown",
        )
        return
    data = await state.get_data()
    master_id = data["schedule_master_id"]
    await db.set_schedule_template(master_id, sh, sm, eh, em)
    await state.clear()
    tmpl = await db.get_schedule_template(master_id)
    masters = await db.get_all_masters()
    master = next((m for m in masters if m["id"] == master_id), None)
    name = master["name"] if master else "Мастер"
    await message.answer(
        f"✅ Шаблон сохранён для *{name}*: {sh:02d}:{sm:02d} – {eh:02d}:{em:02d}\n\n"
        "Теперь нажмите «✅ Применить расписание», чтобы сгенерировать слоты.",
        parse_mode="Markdown",
        reply_markup=kb_schedule_actions(master_id, tmpl),
    )


@dp.callback_query(F.data.startswith("adms_apply:"))
async def adms_apply_schedule(callback: types.CallbackQuery):
    if not is_admin(callback.from_user.id):
        return
    master_id = int(callback.data.split(":")[1])
    tmpl = await db.get_schedule_template(master_id)
    if not tmpl:
        await callback.answer("❌ Сначала задайте часы работы.", show_alert=True)
        return
    await callback.answer("⏳ Генерирую слоты...", show_alert=False)
    added = await db.apply_schedule_template(master_id, days=90)
    masters = await db.get_all_masters()
    master = next((m for m in masters if m["id"] == master_id), None)
    name = master["name"] if master else "Мастер"
    await callback.message.edit_text(
        f"✅ Расписание применено для *{name}*!\n\n"
        f"➕ Добавлено новых слотов: *{added}*\n"
        f"⏰ {tmpl['start_hour']:02d}:{tmpl['start_min']:02d} – "
        f"{tmpl['end_hour']:02d}:{tmpl['end_min']:02d}, каждые {tmpl['interval']} мин\n"
        f"📅 На 90 дней вперёд",
        parse_mode="Markdown",
        reply_markup=kb_schedule_actions(master_id, tmpl),
    )


@dp.callback_query(F.data.startswith("adms_dayoff:"))
async def adms_dayoff_start(callback: types.CallbackQuery):
    if not is_admin(callback.from_user.id):
        return
    master_id = int(callback.data.split(":")[1])
    masters = await db.get_all_masters()
    master = next((m for m in masters if m["id"] == master_id), None)
    name = master["name"] if master else "Мастер"
    await callback.message.edit_text(
        f"🚫 Выберите день выходного для *{name}*:\n_(удалятся только свободные слоты)_",
        parse_mode="Markdown",
        reply_markup=kb_dayoff_dates(master_id),
    )


@dp.callback_query(F.data.startswith("adms_dodate:"))
async def adms_dayoff_confirm(callback: types.CallbackQuery):
    if not is_admin(callback.from_user.id):
        return
    _, master_id_str, date_str = callback.data.split(":", 2)
    master_id = int(master_id_str)
    deleted, kept = await db.delete_day_slots(master_id, date_str)
    d = date.fromisoformat(date_str)
    masters = await db.get_all_masters()
    master = next((m for m in masters if m["id"] == master_id), None)
    name = master["name"] if master else "Мастер"
    tmpl = await db.get_schedule_template(master_id)
    kept_text = f"\n⚠️ *{kept} записей сохранено* (есть клиенты — отмените вручную)" if kept else ""
    await callback.message.edit_text(
        f"✅ Выходной установлен для *{name}*\n"
        f"📅 {d.strftime('%d.%m.%Y')} ({DAYS_RU[d.weekday()]})\n"
        f"🗑 Удалено слотов: {deleted}{kept_text}",
        parse_mode="Markdown",
        reply_markup=kb_schedule_actions(master_id, tmpl),
    )


# ─────────────────────────── Master: auth ─────────────────────

@dp.message(StateFilter(MasterAuth.entering_password))
async def master_auth(message: types.Message, state: FSMContext):
    # BlockedUserMiddleware already handles lockout — no extra check needed here
    user_id = message.from_user.id
    try:
        await message.delete()
    except Exception:
        pass
    if await db.verify_password(user_id, message.text.strip()):
        await db.reset_wrong_attempts(user_id)
        await state.clear()
        master = await db.get_master(user_id)
        await message.answer(
            f"✅ Вход выполнен! Добро пожаловать, *{master['name']}*!",
            parse_mode="Markdown", reply_markup=kb_master_menu(),
        )
    else:
        attempts, lockout_secs = await db.increment_wrong_attempts(user_id)
        if lockout_secs > 0:
            await state.clear()
            await message.answer(
                f"🚫 Слишком много неверных попыток. Аккаунт заблокирован на {db._LOCKOUT_MINUTES} мин."
            )
        else:
            remaining = db._MAX_ATTEMPTS - attempts
            await message.answer(f"❌ Неверный пароль. Осталось попыток: {remaining}")


# ─────────────────────────── Role select ──────────────────────

@dp.callback_query(F.data == "role:master")
async def role_master(callback: types.CallbackQuery, state: FSMContext):
    master = await db.get_master(callback.from_user.id)
    if master:
        if not master["password_hash"]:
            await state.set_state(MasterReg.setting_password)
            await state.update_data(name=master["name"], service_type=master["service_type"])
            await callback.message.edit_text(
                f"🔐 Установите пароль для входа:\n_(минимум 4 символа)_",
                parse_mode="Markdown",
            )
        else:
            await state.set_state(MasterAuth.entering_password)
            await callback.message.edit_text(
                f"🔐 Добро пожаловать, *{master['name']}*!\n\nВведите ваш пароль:\n_(или /cancel для отмены)_",
                parse_mode="Markdown",
            )
        return
    await state.set_state(MasterReg.entering_code)
    await callback.message.edit_text(
        "🔑 Введите код приглашения или код регистрации мастера:\n\n_(или /cancel для возврата в главное меню)_",
    )


@dp.callback_query(F.data == "role:client")
async def role_client(callback: types.CallbackQuery, state: FSMContext):
    await state.set_state(ClientBooking.selecting_service)
    await callback.message.edit_text("🔍 Выберите тип услуги:", reply_markup=kb_services("cs"))


# ─────────────────────────── Master registration ──────────────

@dp.message(StateFilter(MasterReg.entering_code))
async def master_enter_code(message: types.Message, state: FSMContext):
    try:
        await message.delete()
    except Exception:
        pass
    code = message.text.strip()

    # If already a master — redirect to login
    if await db.get_master(message.from_user.id):
        await state.clear()
        await message.answer("✅ Вы уже зарегистрированы как мастер. Нажмите /start для входа.")
        return

    # Check invite code first (admin-created master)
    invite = await db.get_invite(code)
    if invite:
        try:
            await db.use_invite(code, message.from_user.id)
        except Exception:
            # telegram_id already linked to another master account
            await message.answer(
                "⚠️ Этот Telegram-аккаунт уже зарегистрирован как мастер.\n"
                "Используйте /start для входа."
            )
            await state.clear()
            return
        await state.update_data(name=invite["name"], service_type=invite["service_type"])
        await state.set_state(MasterReg.setting_password)
        await message.answer(
            f"✅ Аккаунт подтверждён!\n\n"
            f"👤 {invite['name']} — {SERVICES.get(invite['service_type'])}\n\n"
            "🔐 Придумайте свой пароль для входа:\n_(минимум 4 символа)_",
            parse_mode="Markdown",
        )
        return

    # Otherwise check global MASTER_CODE (self-registration)
    if code != MASTER_CODE:
        await message.answer("❌ Неверный код. Попробуйте ещё раз или /start для возврата.")
        return
    await state.set_state(MasterReg.name)
    await message.answer("✅ Код верный!\n\n✏️ Введите ваше имя (будет видно клиентам):",
                         reply_markup=kb_cancel_text())


@dp.message(StateFilter(MasterReg.name))
async def master_enter_name(message: types.Message, state: FSMContext):
    name = message.text.strip()
    if len(name) < 2:
        await message.answer("❌ Имя слишком короткое:")
        return
    await state.update_data(name=name)
    await state.set_state(MasterReg.service_type)
    await message.answer("🎨 Выберите тип услуги:", reply_markup=kb_services("rs"))


@dp.callback_query(F.data.startswith("rs:"))
async def master_select_service(callback: types.CallbackQuery, state: FSMContext):
    await state.update_data(service_type=callback.data.split(":")[1])
    await state.set_state(MasterReg.setting_password)
    await callback.message.edit_text(
        "🔐 Придумайте пароль для входа:\n_(минимум 4 символа, или /cancel для отмены)_",
        parse_mode="Markdown",
    )


@dp.message(StateFilter(MasterReg.setting_password))
async def master_set_password(message: types.Message, state: FSMContext):
    try:
        await message.delete()
    except Exception:
        pass
    if len(message.text.strip()) < 4:
        await message.answer("❌ Пароль слишком короткий (минимум 4 символа):")
        return
    await state.update_data(new_password=message.text.strip())
    await state.set_state(MasterReg.confirming_password)
    await message.answer("🔐 Повторите пароль:")


@dp.message(StateFilter(MasterReg.confirming_password))
async def master_confirm_password(message: types.Message, state: FSMContext):
    try:
        await message.delete()
    except Exception:
        pass
    data = await state.get_data()
    if message.text.strip() != data["new_password"]:
        await state.set_state(MasterReg.setting_password)
        await message.answer("❌ Пароли не совпадают. Введите пароль заново:")
        return

    # For admin-created masters: just update password (master already exists in DB)
    # For self-registered: insert new record
    existing = await db.get_master(message.from_user.id)
    if existing:
        await db.update_password(message.from_user.id, data["new_password"])
    else:
        await db.register_master(
            message.from_user.id, data["name"], data["service_type"], data["new_password"]
        )
    await state.clear()
    await message.answer(
        f"✅ *Готово!*\n👤 {data['name']} — {SERVICES.get(data['service_type'])}\n🔐 Пароль установлен",
        parse_mode="Markdown", reply_markup=kb_master_menu(),
    )


# ─────────────────────────── Master: add slots ────────────────

@dp.message(F.text == "📅 Добавить слоты")
async def master_add_slots(message: types.Message, state: FSMContext):
    master = await db.get_master(message.from_user.id)
    if not master:
        return
    await state.set_state(MasterSlots.selecting_date)
    await state.update_data(master_id=master["id"], selected_times=[])
    await message.answer("📅 Выберите месяц:", reply_markup=kb_months())


@dp.callback_query(F.data.startswith("slot_month:"), StateFilter(MasterSlots.selecting_date))
async def slot_month_chosen(callback: types.CallbackQuery, state: FSMContext):
    year, month = map(int, callback.data.split(":")[1].split("-"))
    await callback.message.edit_text(
        f"📅 Выберите день — {MONTHS_RU[month]} {year}:",
        reply_markup=kb_days_in_month(year, month),
    )
    await callback.answer()


@dp.callback_query(F.data == "slot_back_months", StateFilter(MasterSlots.selecting_date))
async def slot_back_months(callback: types.CallbackQuery):
    await callback.message.edit_text("📅 Выберите месяц:", reply_markup=kb_months())
    await callback.answer()


@dp.callback_query(F.data.startswith("slot_date:"), StateFilter(MasterSlots.selecting_date))
async def slot_date_chosen(callback: types.CallbackQuery, state: FSMContext):
    sel_date = callback.data.split(":")[1]
    await state.update_data(selected_date=sel_date, selected_times=[])
    await state.set_state(MasterSlots.selecting_times)
    d = date.fromisoformat(sel_date)
    await callback.message.edit_text(
        f"🕐 Время на *{d.strftime('%d.%m.%Y')} ({DAYS_RU[d.weekday()]})*:\nНажимайте — ✅ выбрано",
        parse_mode="Markdown", reply_markup=kb_times([], sel_date),
    )


@dp.callback_query(F.data.startswith("slot_time:"), StateFilter(MasterSlots.selecting_times))
async def slot_time_toggle(callback: types.CallbackQuery, state: FSMContext):
    parts = callback.data.split(":")
    sel_date, time_str = parts[1], f"{parts[2]}:{parts[3]}"
    data = await state.get_data()
    selected: list = data.get("selected_times", [])
    if time_str in selected:
        selected.remove(time_str)
    else:
        selected.append(time_str)
    await state.update_data(selected_times=selected)
    d = date.fromisoformat(sel_date)
    await callback.message.edit_text(
        f"🕐 Время на *{d.strftime('%d.%m.%Y')} ({DAYS_RU[d.weekday()]})*:\nВыбрано: {len(selected)}",
        parse_mode="Markdown", reply_markup=kb_times(selected, sel_date),
    )
    await callback.answer()


@dp.callback_query(F.data.startswith("save_slots:"), StateFilter(MasterSlots.selecting_times))
async def save_slots(callback: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    selected: list = data.get("selected_times", [])
    if not selected:
        await callback.answer("Выберите хотя бы одно время!", show_alert=True)
        return
    added = 0
    for t in selected:
        if await db.add_slot(data["master_id"], data["selected_date"], t):
            added += 1
    await state.clear()
    d = date.fromisoformat(data["selected_date"])
    await callback.message.edit_text(
        f"✅ Добавлено *{added}* слот(ов) на {d.strftime('%d.%m.%Y')} ({DAYS_RU[d.weekday()]}):\n"
        f"{', '.join(sorted(selected))}",
        parse_mode="Markdown",
    )
    await callback.message.answer("Главное меню:", reply_markup=kb_master_menu())


# ─────────────────────────── Master: schedule ─────────────────

@dp.message(F.text == "📋 Мой график")
async def master_schedule(message: types.Message):
    master = await db.get_master(message.from_user.id)
    if not master:
        return
    schedule = await db.get_master_schedule(master["id"], offset_days=0, days=7)
    text, kb = build_schedule_message(schedule, offset=0)
    await message.answer(text, parse_mode="Markdown", reply_markup=kb)


@dp.callback_query(F.data.startswith("sched_page:"))
async def schedule_page(callback: types.CallbackQuery):
    offset = int(callback.data.split(":")[1])
    master = await db.get_master(callback.from_user.id)
    if not master:
        return
    schedule = await db.get_master_schedule(master["id"], offset_days=offset, days=7)
    text, kb = build_schedule_message(schedule, offset=offset)
    await callback.message.edit_text(text, parse_mode="Markdown", reply_markup=kb)
    await callback.answer()


@dp.callback_query(F.data.startswith("clear_day:"))
async def clear_day_slots(callback: types.CallbackQuery):
    date_str = callback.data.split(":", 1)[1]
    master = await db.get_master(callback.from_user.id)
    if not master:
        return
    deleted, kept = await db.delete_day_slots(master["id"], date_str)
    d = date.fromisoformat(date_str)
    msg = f"🗑 Удалено {deleted} слотов за {d.strftime('%d.%m.%Y')}"
    if kept:
        msg += f"\n⚠️ {kept} занятых слотов оставлено"
    await callback.answer(msg, show_alert=True)
    schedule = await db.get_master_schedule(master["id"], offset_days=0, days=7)
    text, kb = build_schedule_message(schedule, offset=0)
    await callback.message.edit_text(text, parse_mode="Markdown", reply_markup=kb)


@dp.callback_query(F.data.startswith("del_slot:"))
async def delete_slot(callback: types.CallbackQuery):
    slot_id = int(callback.data.split(":")[1])
    master = await db.get_master(callback.from_user.id)
    if not master:
        return
    if await db.delete_free_slot(slot_id, master["id"]):
        await callback.answer("✅ Слот удалён.")
        schedule = await db.get_master_schedule(master["id"], offset_days=0, days=7)
        text, kb = build_schedule_message(schedule, offset=0)
        await callback.message.edit_text(text, parse_mode="Markdown", reply_markup=kb)
    else:
        await callback.answer("❌ Слот занят или не найден.", show_alert=True)


@dp.callback_query(F.data.startswith("mcb:"))
async def master_cancel_request(callback: types.CallbackQuery):
    slot_id = int(callback.data.split(":")[1])
    slot = await db.get_slot_info(slot_id)
    if not slot:
        await callback.answer("Запись не найдена.", show_alert=True)
        return
    d = date.fromisoformat(slot["date"])
    await callback.message.answer(
        f"Отменить запись?\n📅 {d.strftime('%d.%m.%Y')} ({DAYS_RU[d.weekday()]}) в {slot['time']}\n"
        "Клиент получит уведомление.",
        reply_markup=kb_cancel_confirm(f"mcb_ok:{slot_id}"),
    )
    await callback.answer()


@dp.callback_query(F.data.startswith("mcb_ok:"))
async def master_cancel_confirm(callback: types.CallbackQuery):
    slot_id = int(callback.data.split(":")[1])
    info = await db.cancel_booking(slot_id, "master", callback.from_user.id)
    if not info:
        await callback.message.edit_text("❌ Запись не найдена или уже отменена.")
        return
    d = date.fromisoformat(info["date"])
    await callback.message.edit_text(
        f"✅ Запись отменена.\n👤 {info['client_name']} — {d.strftime('%d.%m.%Y')} в {info['time']}"
    )
    if info.get("client_telegram_id"):
        try:
            await bot.send_message(
                info["client_telegram_id"],
                f"❌ *Ваша запись отменена мастером*\n\n"
                f"👨‍🎨 {info['master_name']}\n📅 {d.strftime('%d.%m.%Y')} в {info['time']}\n\n"
                "Запишитесь на другое время через /start",
                parse_mode="Markdown",
            )
        except Exception as e:
            logger.warning(f"Не удалось уведомить клиента: {e}")
    master = await db.get_master(callback.from_user.id)
    if master:
        schedule = await db.get_master_schedule(master["id"], offset_days=0, days=7)
        text, kb = build_schedule_message(schedule, offset=0)
        await callback.message.answer(text, parse_mode="Markdown", reply_markup=kb)


@dp.callback_query(F.data == "confirm_no")
async def confirm_no(callback: types.CallbackQuery):
    await callback.message.edit_text("↩️ Отмена действия.")


@dp.callback_query(F.data.startswith("confirm_book:"))
async def confirm_booking_cb(callback: types.CallbackQuery):
    booking_id = int(callback.data.split(":")[1])
    master = await db.get_master(callback.from_user.id)
    if not master:
        await callback.answer("❌ Ошибка авторизации.", show_alert=True)
        return
    result = await db.confirm_booking(booking_id, master["id"])
    if not result:
        await callback.answer("❌ Запрос не найден или уже обработан.", show_alert=True)
        return
    d = date.fromisoformat(result["date"])
    await callback.message.edit_text(
        f"✅ *Запись подтверждена*\n👤 {result['client_name']}\n"
        f"📅 {d.strftime('%d.%m.%Y')} ({DAYS_RU[d.weekday()]}) в {result['time']}",
        parse_mode="Markdown",
    )
    await callback.answer("✅ Подтверждено!")
    if result.get("client_telegram_id"):
        try:
            await bot.send_message(
                result["client_telegram_id"],
                f"✅ *Ваша запись подтверждена!*\n\n"
                f"👨‍🎨 {master['name']}\n"
                f"📅 {d.strftime('%d.%m.%Y')} ({DAYS_RU[d.weekday()]}) в {result['time']}\n\n"
                "Ждём вас! ⏰ _Напоминание придёт за 24ч и за 1ч._\n"
                "❌ _Отменить — /mybookings_",
                parse_mode="Markdown",
            )
        except Exception as e:
            logger.warning(f"Не удалось уведомить клиента о подтверждении: {e}")


@dp.callback_query(F.data.startswith("reject_book:"))
async def reject_booking_cb(callback: types.CallbackQuery):
    booking_id = int(callback.data.split(":")[1])
    master = await db.get_master(callback.from_user.id)
    if not master:
        await callback.answer("❌ Ошибка авторизации.", show_alert=True)
        return
    result = await db.reject_booking(booking_id, master["id"])
    if not result:
        await callback.answer("❌ Запрос не найден или уже обработан.", show_alert=True)
        return
    d = date.fromisoformat(result["date"])
    await callback.message.edit_text(
        f"❌ *Запрос отклонён*\n👤 {result['client_name']}\n"
        f"📅 {d.strftime('%d.%m.%Y')} ({DAYS_RU[d.weekday()]}) в {result['time']}",
        parse_mode="Markdown",
    )
    await callback.answer("❌ Отклонено.")
    if result.get("client_telegram_id"):
        try:
            await bot.send_message(
                result["client_telegram_id"],
                f"❌ *Мастер не может принять вас в это время*\n\n"
                f"👨‍🎨 {master['name']}\n"
                f"📅 {d.strftime('%d.%m.%Y')} ({DAYS_RU[d.weekday()]}) в {result['time']}\n\n"
                "Пожалуйста, выберите другое время через /start → «👤 Я клиент».",
                parse_mode="Markdown",
            )
        except Exception as e:
            logger.warning(f"Не удалось уведомить клиента об отклонении: {e}")


@dp.callback_query(F.data == "noop")
async def noop(callback: types.CallbackQuery):
    await callback.answer()


# ─────────────────────────── Master: history ──────────────────

def kb_history(history) -> InlineKeyboardMarkup:
    b = InlineKeyboardBuilder()
    for h in history:
        d = date.fromisoformat(h["date"])
        if h["cancelled"]:
            icon = "🚫"
            label = f"{icon} {d.strftime('%d.%m')} {h['time']} — {h['client_name']} (отменён)"
            b.button(text=label, callback_data="noop")
        elif h["attended"] == 1:
            icon = "✅"
            label = f"{icon} {d.strftime('%d.%m')} {h['time']} — {h['client_name']}"
            b.button(text=label, callback_data="noop")
            b.button(text="❌ Не пришёл", callback_data=f"att:2:{h['booking_id']}")
            b.button(text="🗑 Удалить", callback_data=f"hdel:{h['slot_id']}")
        elif h["attended"] == 2:
            icon = "❌"
            label = f"{icon} {d.strftime('%d.%m')} {h['time']} — {h['client_name']}"
            b.button(text=label, callback_data="noop")
            b.button(text="✅ Пришёл", callback_data=f"att:1:{h['booking_id']}")
            b.button(text="🗑 Удалить", callback_data=f"hdel:{h['slot_id']}")
        else:
            label = f"⬜ {d.strftime('%d.%m')} {h['time']} — {h['client_name']}"
            b.button(text=label, callback_data="noop")
            b.button(text="✅ Пришёл",  callback_data=f"att:1:{h['booking_id']}")
            b.button(text="❌ Не пришёл", callback_data=f"att:2:{h['booking_id']}")
            b.button(text="🗑 Удалить",  callback_data=f"hdel:{h['slot_id']}")
    b.adjust(1)
    return b.as_markup()


@dp.message(F.text == "🕐 История")
async def master_history(message: types.Message):
    master = await db.get_master(message.from_user.id)
    if not master:
        return
    history = await db.get_master_history(master["id"], days=30)
    if not history:
        await message.answer("📭 За последние 30 дней записей нет.")
        return
    await message.answer(
        "🕐 *История за 30 дней:*\n\n"
        "⬜ — статус не отмечен\n✅ — пришёл · ❌ — не пришёл · 🚫 — отменён",
        parse_mode="Markdown",
        reply_markup=kb_history(history),
    )


@dp.callback_query(F.data.startswith("att:"))
async def mark_attendance(callback: types.CallbackQuery):
    parts = callback.data.split(":")
    status, booking_id = int(parts[1]), int(parts[2])
    await db.mark_attended(booking_id, status)
    label = "✅ Отмечен как пришедший" if status == 1 else "❌ Отмечен как не пришедший"
    await callback.answer(label, show_alert=False)
    master = await db.get_master(callback.from_user.id)
    if master:
        history = await db.get_master_history(master["id"], days=30)
        if history:
            await callback.message.edit_reply_markup(reply_markup=kb_history(history))


@dp.callback_query(F.data.startswith("hdel:"))
async def history_delete(callback: types.CallbackQuery):
    slot_id = int(callback.data.split(":")[1])
    master = await db.get_master(callback.from_user.id)
    if not master:
        return
    result = await db.master_delete_booking(slot_id, master["id"])
    if result:
        await callback.answer("🗑 Запись удалена.", show_alert=False)
        history = await db.get_master_history(master["id"], days=30)
        if history:
            await callback.message.edit_reply_markup(reply_markup=kb_history(history))
        else:
            await callback.message.edit_text("📭 История пуста.")
    else:
        await callback.answer("❌ Не найдено или уже удалено.", show_alert=True)


# Кнопка «🗑 Удалить запись» в графике (клиент отменил лично/по телефону)
@dp.callback_query(F.data.startswith("manualdel:"))
async def manual_delete_upcoming(callback: types.CallbackQuery):
    slot_id = int(callback.data.split(":")[1])
    master = await db.get_master(callback.from_user.id)
    if not master:
        return
    result = await db.master_delete_booking(slot_id, master["id"])
    if result:
        await callback.answer("🗑 Запись удалена, слот освобождён.", show_alert=False)
        schedule = await db.get_master_schedule(master["id"], offset_days=0, days=7)
        text, kb = build_schedule_message(schedule, offset=0)
        await callback.message.edit_text(text, parse_mode="Markdown", reply_markup=kb)
    else:
        await callback.answer("❌ Запись не найдена.", show_alert=True)


# ─────────────────────────── Master: Excel export ─────────────

@dp.message(F.text == "📥 Экспорт Excel")
async def master_export_excel(message: types.Message, master=None):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if master is None:
        master = await db.get_master(message.from_user.id)
    if not master:
        return

    bookings = await db.get_all_master_bookings(master["id"])
    if not bookings:
        await message.answer("📭 Записей пока нет — экспортировать нечего.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Клиенты"

    # ── Styles ──────────────────────────────────────────────────
    header_font    = Font(bold=True, color="FFFFFF", size=11)
    header_fill    = PatternFill("solid", fgColor="6A0DAD")   # purple
    ok_fill        = PatternFill("solid", fgColor="D4EDDA")   # green tint
    cancel_fill    = PatternFill("solid", fgColor="F8D7DA")   # red tint
    center         = Alignment(horizontal="center", vertical="center")
    thin           = Side(style="thin", color="BBBBBB")
    border         = Border(left=thin, right=thin, top=thin, bottom=thin)

    ATTENDED_LABEL = {"0": "—", "1": "✅ Пришёл", "2": "❌ Не пришёл"}

    # ── Header ──────────────────────────────────────────────────
    headers = ["№", "Имя клиента", "Телефон", "Дата", "Время",
               "Услуга", "Посещение", "Статус"]
    col_widths = [5, 22, 16, 12, 8, 16, 14, 14]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22

    # ── Data rows ───────────────────────────────────────────────
    svc_label = SERVICES.get(master["service_type"], master["service_type"])
    for i, b in enumerate(bookings, 1):
        cancelled  = bool(b["cancelled"])
        attended   = str(b.get("attended", 0))
        status_txt = "Отменён" if cancelled else "Активна"
        row_fill   = cancel_fill if cancelled else ok_fill

        row_data = [
            i,
            b["client_name"],
            b["client_phone"],
            b["date"],
            b["time"],
            svc_label,
            ATTENDED_LABEL.get(attended, "—"),
            status_txt,
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            cell.alignment = center
            cell.border    = border
            cell.fill      = row_fill

    # ── Freeze header, auto-filter ───────────────────────────────
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{len(bookings) + 1}"

    # ── Summary block ────────────────────────────────────────────
    total     = len(bookings)
    active    = sum(1 for b in bookings if not b["cancelled"])
    cancelled = total - active
    attended  = sum(1 for b in bookings if b.get("attended") == 1)

    summary_row = len(bookings) + 3
    ws.cell(row=summary_row,     column=1, value="Итого записей:").font  = Font(bold=True)
    ws.cell(row=summary_row,     column=2, value=total)
    ws.cell(row=summary_row + 1, column=1, value="Активных:").font       = Font(bold=True)
    ws.cell(row=summary_row + 1, column=2, value=active)
    ws.cell(row=summary_row + 2, column=1, value="Отменено:").font       = Font(bold=True)
    ws.cell(row=summary_row + 2, column=2, value=cancelled)
    ws.cell(row=summary_row + 3, column=1, value="Пришло:").font         = Font(bold=True)
    ws.cell(row=summary_row + 3, column=2, value=attended)

    # ── Save to buffer and send ──────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from aiogram.types import BufferedInputFile
    today_str = date.today().strftime("%d.%m.%Y")
    filename  = f"clients_{master['name']}_{today_str}.xlsx".replace(" ", "_")

    await message.answer_document(
        BufferedInputFile(buf.read(), filename=filename),
        caption=(
            f"📥 *Экспорт клиентов — {master['name']}*\n"
            f"📅 {today_str}\n\n"
            f"📋 Всего: {total} · ✅ Активных: {active} · "
            f"❌ Отменено: {cancelled} · 🟢 Пришло: {attended}"
        ),
        parse_mode="Markdown",
    )


# ─────────────────────────── Master: clients ──────────────────

@dp.message(F.text == "👥 Клиенты")
async def master_clients(message: types.Message):
    master = await db.get_master(message.from_user.id)
    if not master:
        return
    clients = await db.get_master_clients(master["id"])
    if not clients:
        await message.answer("👥 Клиентов пока нет.")
        return
    text = "👥 *Ваши клиенты:*\n\n"
    for i, c in enumerate(clients, 1):
        d = date.fromisoformat(c["date"])
        text += (
            f"*{i}. {c['client_name']}*\n"
            f"   📱 {c['client_phone']}\n"
            f"   📅 {d.strftime('%d.%m.%Y')} ({DAYS_RU[d.weekday()]}) в {c['time']}\n\n"
        )
    await message.answer(text, parse_mode="Markdown")


# ─────────────────────────── Master: stats ────────────────────

@dp.message(F.text == "📊 Статистика")
async def master_stats(message: types.Message):
    master = await db.get_master(message.from_user.id)
    if not master:
        return
    s = await db.get_master_stats(master["id"])
    await message.answer(
        f"📊 *Статистика*\n\n"
        f"*{date.today().strftime('%B %Y')}:*\n"
        f"✅ Записей: {s['this_month']}\n❌ Отменено: {s['cancelled_month']}\n\n"
        f"*За всё время:*\n👥 Клиентов: {s['unique_clients']}\n📋 Записей: {s['total_all']}",
        parse_mode="Markdown",
    )


# ─────────────────────────── Master: broadcast ────────────────

@dp.message(F.text == "📢 Рассылка")
async def master_broadcast_start(message: types.Message, state: FSMContext):
    master = await db.get_master(message.from_user.id)
    if not master:
        return
    ids = await db.get_broadcast_client_ids(master["id"])
    if not ids:
        await message.answer("📢 Нет клиентов для рассылки.")
        return
    await state.set_state(MasterBroadcast.entering_text)
    await state.update_data(master_id=master["id"], client_ids=ids)
    await message.answer(f"📢 Рассылка по *{len(ids)}* клиентам.\n\nВведите текст сообщения:",
                         parse_mode="Markdown", reply_markup=kb_cancel_text())


@dp.message(StateFilter(MasterBroadcast.entering_text))
async def master_broadcast_text(message: types.Message, state: FSMContext):
    await state.update_data(broadcast_text=message.text.strip())
    b = InlineKeyboardBuilder()
    b.button(text="📤 Отправить", callback_data="broadcast_send")
    b.button(text="❌ Отмена",    callback_data="broadcast_cancel")
    b.adjust(2)
    await message.answer(
        f"📋 *Предпросмотр:*\n\n{message.text.strip()}\n\nОтправить всем клиентам?",
        parse_mode="Markdown", reply_markup=b.as_markup(),
    )


@dp.callback_query(F.data == "broadcast_send")
async def broadcast_send(callback: types.CallbackQuery, state: FSMContext):
    if not await db.get_master(callback.from_user.id):
        await callback.answer("❌ Доступ запрещён.", show_alert=True)
        return
    data = await state.get_data()
    text = data.get("broadcast_text", "")
    ids: list = data.get("client_ids", [])
    await state.clear()
    await callback.message.edit_text("📤 Отправляю...")
    sent = 0
    for tg_id in ids:
        try:
            await bot.send_message(tg_id, text)
            sent += 1
        except Exception:
            pass
    await callback.message.edit_text(f"✅ Рассылка завершена! Отправлено: {sent}/{len(ids)}")
    await callback.message.answer("Главное меню:", reply_markup=kb_master_menu())


@dp.callback_query(F.data == "broadcast_cancel")
async def broadcast_cancel(callback: types.CallbackQuery, state: FSMContext):
    if not await db.get_master(callback.from_user.id):
        await callback.answer("❌ Доступ запрещён.", show_alert=True)
        return
    await state.clear()
    await callback.message.edit_text("❌ Рассылка отменена.")
    await callback.message.answer("Главное меню:", reply_markup=kb_master_menu())


# ─────────────────────────── Master: settings ─────────────────

@dp.message(F.text == "⚙️ Настройки")
async def master_settings(message: types.Message):
    master = await db.get_master(message.from_user.id)
    price_label = f"{master['price']} ₽" if master and master["price"] else "не указана"
    b = InlineKeyboardBuilder()
    b.button(text=f"💰 Цена услуги ({price_label})", callback_data="set_price")
    b.button(text="🔑 Изменить пароль",               callback_data="change_pwd")
    b.button(text="🔄 Изменить имя / услугу",         callback_data="reset_master")
    b.adjust(1)
    await message.answer("⚙️ Настройки:", reply_markup=b.as_markup())


@dp.callback_query(F.data == "set_price")
async def set_price_start(callback: types.CallbackQuery, state: FSMContext):
    if not await db.get_master(callback.from_user.id):
        await callback.answer("❌ Доступ запрещён.", show_alert=True)
        return
    await state.set_state(MasterPriceSetting.entering_price)
    await callback.message.edit_text("💰 Введите стоимость услуги в рублях (0 — убрать цену):\n_(или /cancel для отмены)_")


@dp.message(StateFilter(MasterPriceSetting.entering_price))
async def set_price_enter(message: types.Message, state: FSMContext):
    if not message.text.strip().isdigit():
        await message.answer("❌ Введите число (например: 1500):")
        return
    await db.set_master_price(message.from_user.id, int(message.text.strip()))
    await state.clear()
    await message.answer("✅ Цена обновлена!", reply_markup=kb_master_menu())


@dp.callback_query(F.data == "reset_master")
async def reset_master(callback: types.CallbackQuery, state: FSMContext):
    if not await db.get_master(callback.from_user.id):
        await callback.answer("❌ Доступ запрещён.", show_alert=True)
        return
    await state.set_state(MasterUpdateInfo.name)
    await callback.message.edit_text("✏️ Введите новое имя (будет видно клиентам):\n_(или /cancel для отмены)_")


@dp.message(StateFilter(MasterUpdateInfo.name))
async def update_info_name(message: types.Message, state: FSMContext):
    if not await db.get_master(message.from_user.id):
        return
    name = message.text.strip()
    if len(name) < 2:
        await message.answer("❌ Имя слишком короткое:")
        return
    await state.update_data(new_name=name)
    await state.set_state(MasterUpdateInfo.service_type)
    await message.answer("🎨 Выберите новый тип услуги:", reply_markup=kb_services("upd"))


@dp.callback_query(F.data.startswith("upd:"), StateFilter(MasterUpdateInfo.service_type))
async def update_info_service(callback: types.CallbackQuery, state: FSMContext):
    service_type = callback.data.split(":")[1]
    data = await state.get_data()
    name = data["new_name"]
    await db.register_master(callback.from_user.id, name, service_type,
                             password=None)  # update name/service only
    await state.clear()
    await callback.message.edit_text(
        f"✅ Данные обновлены!\n👤 {name} — {SERVICES.get(service_type)}",
        parse_mode="Markdown",
        reply_markup=None,
    )
    await callback.message.answer("Главное меню:", reply_markup=kb_master_menu())


@dp.callback_query(F.data == "change_pwd")
async def change_pwd_start(callback: types.CallbackQuery, state: FSMContext):
    if not await db.get_master(callback.from_user.id):
        await callback.answer("❌ Доступ запрещён.", show_alert=True)
        return
    await state.set_state(MasterChangePwd.old_password)
    await callback.message.edit_text("🔑 Введите текущий пароль:\n_(или /cancel для отмены)_")


@dp.message(StateFilter(MasterChangePwd.old_password))
async def change_pwd_old(message: types.Message, state: FSMContext):
    try:
        await message.delete()
    except Exception:
        pass
    if not await db.verify_password(message.from_user.id, message.text.strip()):
        await message.answer("❌ Неверный пароль:")
        return
    await state.set_state(MasterChangePwd.new_password)
    await message.answer("🔐 Введите новый пароль (минимум 4 символа):")


@dp.message(StateFilter(MasterChangePwd.new_password))
async def change_pwd_new(message: types.Message, state: FSMContext):
    try:
        await message.delete()
    except Exception:
        pass
    if len(message.text.strip()) < 4:
        await message.answer("❌ Слишком короткий (минимум 4 символа):")
        return
    await state.update_data(new_password=message.text.strip())
    await state.set_state(MasterChangePwd.confirm_password)
    await message.answer("🔐 Повторите новый пароль:")


@dp.message(StateFilter(MasterChangePwd.confirm_password))
async def change_pwd_confirm(message: types.Message, state: FSMContext):
    try:
        await message.delete()
    except Exception:
        pass
    data = await state.get_data()
    if message.text.strip() != data["new_password"]:
        await state.set_state(MasterChangePwd.new_password)
        await message.answer("❌ Не совпадают. Введите новый пароль заново:")
        return
    await db.update_password(message.from_user.id, data["new_password"])
    await state.clear()
    await message.answer("✅ Пароль изменён!", reply_markup=kb_master_menu())


# ─────────────────────────── Client: my bookings ──────────────

@dp.callback_query(F.data == "my_bookings")
async def client_my_bookings(callback: types.CallbackQuery):
    bookings = await db.get_client_bookings(callback.from_user.id)
    if not bookings:
        await callback.message.edit_text(
            "📋 У вас нет предстоящих записей.\n\nЗапишитесь через «Я клиент».",
            reply_markup=kb_start(),
        )
        return
    now = datetime.now()
    text = "📋 *Ваши предстоящие записи:*\n\n"
    has_cancel = False
    for i, bk in enumerate(bookings, 1):
        d = date.fromisoformat(bk["date"])
        appt_dt = datetime.strptime(f"{bk['date']} {bk['time']}", "%Y-%m-%d %H:%M")
        status_tag = " ⏳ _ожидает подтв._" if bk.get("pending") else ""
        text += (
            f"*{i}.* 📅 {d.strftime('%d.%m.%Y')} ({DAYS_RU[d.weekday()]}) в {bk['time']}{status_tag}\n"
            f"   👨‍🎨 {bk['master_name']} — {SERVICES.get(bk['service_type'], '')}\n\n"
        )
        if (appt_dt - now).total_seconds() / 60 > 120:
            has_cancel = True
    if has_cancel:
        text += "_Кнопки ниже — отмена записи (доступна за 2+ часа):_"
    await callback.message.edit_text(
        text, parse_mode="Markdown", reply_markup=kb_client_bookings(bookings)
    )


@dp.callback_query(F.data.startswith("ccb:"))
async def client_cancel_request(callback: types.CallbackQuery):
    slot_id = int(callback.data.split(":")[1])
    slot = await db.get_slot_info(slot_id)
    if not slot:
        await callback.answer("Запись не найдена.", show_alert=True)
        return
    appt_dt = datetime.strptime(f"{slot['date']} {slot['time']}", "%Y-%m-%d %H:%M")
    if (appt_dt - datetime.now()).total_seconds() < 7200:
        await callback.answer("❌ Отмена недоступна менее чем за 2 часа.", show_alert=True)
        return
    d = date.fromisoformat(slot["date"])
    await callback.message.answer(
        f"Отменить вашу запись?\n\n"
        f"👨‍🎨 {slot['master_name']} — {SERVICES.get(slot['service_type'], '')}\n"
        f"📅 {d.strftime('%d.%m.%Y')} ({DAYS_RU[d.weekday()]}) в {slot['time']}",
        reply_markup=kb_cancel_confirm(f"ccb_ok:{slot_id}"),
    )
    await callback.answer()


@dp.callback_query(F.data.startswith("ccb_ok:"))
async def client_cancel_confirm(callback: types.CallbackQuery):
    slot_id = int(callback.data.split(":")[1])
    info = await db.cancel_booking(slot_id, "client", callback.from_user.id)
    if not info:
        await callback.message.edit_text("❌ Запись не найдена или уже отменена.")
        return
    d = date.fromisoformat(info["date"])
    b = InlineKeyboardBuilder()
    b.button(text="🏠 Главное меню", callback_data="main_menu")
    await callback.message.edit_text(
        f"✅ Запись отменена.\n📅 {d.strftime('%d.%m.%Y')} в {info['time']} — {info['master_name']}",
        reply_markup=b.as_markup(),
    )
    try:
        await bot.send_message(
            info["master_tg_id"],
            f"❌ *Клиент отменил запись*\n\n"
            f"👤 {info['client_name']} — {info['client_phone']}\n"
            f"📅 {d.strftime('%d.%m.%Y')} в {info['time']}",
            parse_mode="Markdown",
        )
    except Exception as e:
        logger.warning(f"Не удалось уведомить мастера: {e}")


# ─────────────────────────── Client: service → book ──────────

@dp.callback_query(F.data.startswith("cs:"), StateFilter(ClientBooking.selecting_service))
async def client_service(callback: types.CallbackQuery, state: FSMContext):
    service_type = callback.data.split(":")[1]
    await state.update_data(service_type=service_type)
    masters = await db.get_masters_by_service(service_type)
    if not masters:
        await callback.message.edit_text(
            f"😔 Мастеров по услуге «{SERVICES.get(service_type)}» пока нет.\n\nВыберите другую:",
            reply_markup=kb_services("cs"),
        )
        return
    b = InlineKeyboardBuilder()
    for m in masters:
        price_str = f" · от {m['price']} ₽" if m["price"] else ""
        b.button(
            text=f"👨‍🎨 {m['name']}{price_str}",
            callback_data=f"cm:{m['id']}:{m['telegram_id']}:{m['price']}",
        )
    b.button(text="🔙 Назад", callback_data="role:client")
    b.adjust(1)
    await state.set_state(ClientBooking.selecting_master)
    await callback.message.edit_text(
        f"Выберите мастера — {SERVICES.get(service_type)}:", reply_markup=b.as_markup()
    )


@dp.callback_query(F.data.startswith("cm:"), StateFilter(ClientBooking.selecting_master))
async def client_master(callback: types.CallbackQuery, state: FSMContext):
    parts = callback.data.split(":")
    master_id, master_tg_id, price = int(parts[1]), int(parts[2]), int(parts[3])
    await state.update_data(master_id=master_id, master_tg_id=master_tg_id, master_price=price)
    await state.set_state(ClientBooking.entering_name)
    await callback.message.edit_text("✏️ Введите ваше имя:\n_(или /cancel для отмены)_")


@dp.message(StateFilter(ClientBooking.entering_name))
async def client_enter_name(message: types.Message, state: FSMContext):
    name = message.text.strip()
    if len(name) < 2:
        await message.answer("❌ Имя слишком короткое:")
        return
    await state.update_data(client_name=name)
    await state.set_state(ClientBooking.entering_phone)
    await message.answer(
        "📱 Введите номер телефона или отправьте контакт:",
        reply_markup=ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text="📱 Отправить мой номер", request_contact=True)],
                [KeyboardButton(text="❌ Отмена")],
            ],
            resize_keyboard=True,
        ),
    )


@dp.message(StateFilter(ClientBooking.entering_phone), F.contact)
async def client_phone_contact(message: types.Message, state: FSMContext):
    await _process_phone(message, state, message.contact.phone_number)


@dp.message(StateFilter(ClientBooking.entering_phone), F.text)
async def client_phone_text(message: types.Message, state: FSMContext):
    phone = message.text.strip()
    if len([c for c in phone if c.isdigit()]) < 7:
        await message.answer("❌ Введите корректный номер (минимум 7 цифр):")
        return
    await _process_phone(message, state, phone)


async def _process_phone(message: types.Message, state: FSMContext, phone: str):
    await state.update_data(client_phone=phone)
    data = await state.get_data()
    slots = await db.get_available_slots(data["master_id"])
    if not slots:
        await message.answer("😔 У мастера нет свободных слотов. Попробуйте позже.",
                             reply_markup=ReplyKeyboardRemove())
        await state.clear()
        return
    await state.set_state(ClientBooking.selecting_slot)
    await message.answer("✅ Данные приняты!", reply_markup=ReplyKeyboardRemove())
    await message.answer("📅 Выберите удобное время:",
                         reply_markup=kb_slots(slots, data.get("master_price", 0)))


@dp.callback_query(F.data.startswith("book:"), StateFilter(ClientBooking.selecting_slot))
async def client_book(callback: types.CallbackQuery, state: FSMContext):
    slot_id = int(callback.data.split(":")[1])
    data = await state.get_data()
    booking_id = await db.book_slot(
        slot_id, data["client_name"], data["client_phone"], callback.from_user.id
    )
    if booking_id is None:
        await callback.answer("❌ Слот уже занят! Выберите другой.", show_alert=True)
        slots = await db.get_available_slots(data["master_id"])
        if slots:
            await callback.message.edit_reply_markup(
                reply_markup=kb_slots(slots, data.get("master_price", 0))
            )
        else:
            await callback.message.edit_text("😔 Свободных слотов больше нет.")
            await state.clear()
        return
    slot = await db.get_slot_info(slot_id)
    d = date.fromisoformat(slot["date"])
    price = data.get("master_price", 0)
    price_line = f"💰 {price} ₽\n" if price else ""
    await state.clear()
    b = InlineKeyboardBuilder()
    b.button(text="📋 Мои записи", callback_data="my_bookings")
    b.button(text="🏠 Главное меню", callback_data="main_menu")
    b.adjust(2)
    await callback.message.edit_text(
        f"⏳ *Запрос отправлен мастеру!*\n\n👨‍🎨 {slot['master_name']}\n"
        f"📅 {d.strftime('%d.%m.%Y')} ({DAYS_RU[d.weekday()]})\n🕐 {slot['time']}\n"
        f"{price_line}\n"
        "Ожидайте подтверждения — вы получите уведомление.\n\n"
        "❌ _Чтобы отменить — /mybookings_",
        parse_mode="Markdown", reply_markup=b.as_markup(),
    )
    try:
        b_confirm = InlineKeyboardBuilder()
        b_confirm.button(text="✅ Подтвердить", callback_data=f"confirm_book:{booking_id}")
        b_confirm.button(text="❌ Отклонить",   callback_data=f"reject_book:{booking_id}")
        b_confirm.adjust(2)
        await bot.send_message(
            data["master_tg_id"],
            f"🔔 *Новый запрос на запись!*\n\n👤 {data['client_name']}\n📱 {data['client_phone']}\n"
            f"📅 {d.strftime('%d.%m.%Y')} ({DAYS_RU[d.weekday()]}) в {slot['time']}\n\n"
            "Подтвердите или отклоните:",
            parse_mode="Markdown",
            reply_markup=b_confirm.as_markup(),
        )
    except Exception as e:
        logger.warning(f"Не удалось уведомить мастера: {e}")


# ─────────────────────────── Reviews ─────────────────────────

@dp.callback_query(F.data.startswith("review:"))
async def save_review_cb(callback: types.CallbackQuery):
    parts = callback.data.split(":")
    booking_id, score = int(parts[1]), int(parts[2])
    await db.save_review(booking_id, score)
    stars = "⭐" * score
    await callback.message.edit_text(f"Спасибо за оценку! {stars}\n\nДо встречи снова! 😊")
    await callback.answer()


# ─────────────────────────── Master: apply template ───────────

@dp.message(F.text == "🔄 Шаблон")
async def master_apply_template_btn(message: types.Message, master=None):
    if master is None:
        master = await db.get_master(message.from_user.id)
    if not master:
        return
    tmpl = await db.get_schedule_template(master["id"])
    if not tmpl:
        await message.answer(
            "ℹ️ Шаблон расписания не настроен.\n\n"
            "Обратитесь к администратору — он задаёт часы работы через /admin → «📅 Расписание мастеров».",
        )
        return
    added = await db.apply_schedule_template(master["id"], days=90)
    await message.answer(
        f"✅ *Расписание обновлено!*\n\n"
        f"⏰ {tmpl['start_hour']:02d}:{tmpl['start_min']:02d} – "
        f"{tmpl['end_hour']:02d}:{tmpl['end_min']:02d}, каждые {tmpl['interval']} мин\n"
        f"➕ Добавлено новых слотов: *{added}*\n"
        f"📅 На 90 дней вперёд",
        parse_mode="Markdown",
        reply_markup=kb_master_menu(),
    )


# ─────────────────────────── Cancel ───────────────────────────

@dp.callback_query(F.data == "main_menu")
async def main_menu(callback: types.CallbackQuery, state: FSMContext):
    await state.clear()
    text = "👋 Добро пожаловать!\n\nВы мастер или клиент?"
    if is_admin(callback.from_user.id):
        text += "\n\n🔑 Вы вошли как *администратор*. Используйте /admin"
    await callback.message.edit_text(text, parse_mode="Markdown", reply_markup=kb_start())


@dp.callback_query(F.data == "cancel")
async def cancel_action(callback: types.CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.edit_text("❌ Действие отменено.\n\nВозврат в главное меню:",
                                     reply_markup=kb_start())


# ─────────────────────────── Main ─────────────────────────────

async def main():
    await storage._init()
    await db.init_db()
    def _on_task_done(name: str):
        def cb(t):
            if not t.cancelled() and t.exception():
                logger.error(f"{name} died: {t.exception()}")
        return cb

    task = asyncio.create_task(reminder_loop())
    task.add_done_callback(_on_task_done("reminder_loop"))
    ctask = asyncio.create_task(cleanup_loop())
    ctask.add_done_callback(_on_task_done("cleanup_loop"))
    if ADMIN_ID == 0:
        logger.warning(
            "ADMIN_ID не установлен в .env — команда /admin недоступна. "
            "Запустите бота, напишите /myid чтобы узнать свой ID, затем добавьте ADMIN_ID=... в .env"
        )
    logger.info("Бот запущен...")
    await dp.start_polling(bot, allowed_updates=dp.resolve_used_update_types())


if __name__ == "__main__":
    asyncio.run(main())
